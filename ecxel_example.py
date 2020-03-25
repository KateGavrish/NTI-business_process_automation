import openpyxl
from sql.data import db_session
from sql.data import __all_models


def import_details(name):
    wb = openpyxl.load_workbook(filename=name)
    sheet = wb['Лист1']

    val = sheet['A1'].value

    vals = []
    n = 0
    while True:
        n += 1
        val = [sheet[f'A{n}'].value, sheet[f'B{n}'].value, sheet[f'C{n}'].value]
        vals.append(val)
        if val[0] is None:
            break
    vals = vals[1:]

    db_session.global_init('sql/db/drons1.sqlite')
    for i in range(len(vals)):
        d = __all_models.details.DetailCategory()
        if vals[i][1] is None or vals[i][2] is None:
            continue
        d.name_det = vals[i][1]
        d.category = vals[i][2]
        session = db_session.create_session()
        session.add(d)
        session.commit()
    return 'ok'


def import_drons(name):
    wb = openpyxl.load_workbook(filename=name)
    sheet = wb['Лист1']

    val = sheet['A1'].value

    vals = []
    n = 0
    while True:
        n += 1
        val = [sheet[f'A{n}'].value, sheet[f'B{n}'].value, sheet[f'C{n}'].value]
        vals.append(val)
        if val[0] is None:
            break
    vals = vals[1:]

    db_session.global_init('sql/db/drons1.sqlite')
    for i in range(len(vals)):
        d = __all_models.drons_cost.DronCost()
        if vals[i][1] is None or vals[i][2] is None or type(vals[i][2]) != int:
            continue
        d.name_dron = vals[i][1]
        d.cost = vals[i][2]
        session = db_session.create_session()
        session.add(d)
        session.commit()
    return 'ok'


def import_technial_map(name):
    wb = openpyxl.load_workbook(filename=name)
    sheet = wb['Лист1']

    val = sheet['A1'].value

    vals = []
    n = 0
    while True:
        n += 1
        val = [sheet[f'A{n}'].value, sheet[f'B{n}'].value, sheet[f'C{n}'].value, sheet[f'D{n}'].value]
        vals.append(val)
        if val[1] is None:
            break
    vals = vals[1:]

    db_session.global_init('sql/db/drons1.sqlite')
    for i in range(len(vals)):
        d = __all_models.count_for_dron.QuanForDron()
        if vals[i][1] is None or vals[i][2] is None or vals[i][3] is None or type(vals[i][3]) != int:
            continue
        d.name_dron = vals[i][1]
        d.name_det = vals[i][2]
        d.quantity = vals[i][3]
        session = db_session.create_session()
        session.add(d)
        session.commit()
    return 'ok'


